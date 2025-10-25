from __future__ import annotations

import json
import math
import os
from datetime import datetime, time as dt_time
from pathlib import Path
from typing import List, Dict, Any, Tuple

from flask import Flask, jsonify, render_template, request, redirect, url_for, session


app = Flask(__name__)


def create_pdf_header():
    """PDF Header Design Code"""
    
    # Get styles
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    styles = getSampleStyleSheet()
    
    # Define header styles
    header_title = ParagraphStyle(
        'HeaderTitle', 
        parent=styles['Title'], 
        alignment=0, 
        fontSize=16, 
        leading=19
    )
    header_sub = ParagraphStyle(
        'HeaderSub', 
        parent=styles['Normal'], 
        alignment=0, 
        fontSize=10, 
        leading=12
    )
    
    # Logo setup
    try:
        from flask import current_app
        from pathlib import Path
        logo_path = Path(__file__).parent / "static" / "images" / "logo-removebg-preview.png"
        if logo_path.exists():
            from reportlab.platypus import Image
            from reportlab.lib.units import mm
            logo_img = Image(str(logo_path))
            logo_img._restrictSize(26*mm, 26*mm)
        else:
            logo_img = ''
    except Exception:
        logo_img = ''
    
    # Header text
    from reportlab.platypus import Paragraph
    header_text = [
        Paragraph('Dr. B. B. Hegde First Grade College, Kundapura', header_title),
        Paragraph('A Unit of Coondapur Education Society (R)', header_sub)
    ]
    
    # Create header table
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    
    header_table = Table(
        [[logo_img, header_text]], 
        colWidths=[26*mm, (A4[0] - (18*mm + 18*mm) - 26*mm)]
    )
    
    # Header table styling
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LINEBELOW', (0,0), (-1,0), 0.75, colors.lightgrey),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING', (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    
    return header_table
# Simple secret key for session management; replace via ENV in production
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-change-me")


# Configure via environment variable or default path; can be overridden at runtime by querystring
ATTENDANCE_DIR = os.environ.get("ATTENDANCE_DIR", r"F:\\Graahi Reports_new\\Graahi Reports\\attendance")


def get_attendance_dir() -> Path:
    # Allow overriding via query parameter for flexibility (read-only)
    override = request.args.get("dir")
    path_str = override if override else ATTENDANCE_DIR
    return Path(path_str)


def read_attendance_for_date(base_dir: Path, date_str: str) -> List[Dict[str, Any]]:
    """Read a JSON file named YYYY-MM-DD.json and return list of rows.

    This is read-only and should not interfere with other processes writing files.
    """
    # Normalize expected filename
    filename = f"{date_str}.json"
    file_path = base_dir / filename
    if not file_path.exists():
        return []
    try:
        with file_path.open("r", encoding="utf-8") as f:  # read-only
            data = json.load(f)
            # Ensure we return a list of dicts
            if isinstance(data, list):
                return data
            return []
    except Exception:
        # Fail closed on parse errors
        return []


def is_logged_in() -> bool:
    return bool(session.get("logged_in"))




def get_attendance_from_source(date_str: str) -> List[Dict[str, Any]]:
    """Read attendance data directly from source directory."""
    source_dir = get_attendance_dir()
    source_file = source_dir / f"{date_str}.json"
    if not source_file.exists():
        return []
    
    try:
        with source_file.open("r", encoding="utf-8") as f:
            data = json.load(f)
            if isinstance(data, list):
                return data
            return []
    except Exception:
        return []


# ---------- Delay annotation logic ----------
FACULTY_DETAIL_PATH = Path.cwd() / "JSON" / "faculty_detail.json"

_faculty_cache: Dict[str, Dict[str, Any]] | None = None

def load_faculty_details() -> Dict[str, Dict[str, Any]]:
    global _faculty_cache
    if _faculty_cache is not None:
        return _faculty_cache
    try:
        with FACULTY_DETAIL_PATH.open("r", encoding="utf-8") as f:
            data = json.load(f)
            if isinstance(data, dict):
                _faculty_cache = data
                return data
    except Exception:
        pass
    _faculty_cache = {}
    return _faculty_cache

def determine_staff_type(faculty_id: str) -> str:
    """Return 'teaching', 'admin', or 'support' using faculty_detail.json.
    - category == 'Teaching Faculty' => teaching
    - category == 'Non-Teaching Faculty' and department includes 'Administrative' => admin
    - department == 'Computer Applications' => support (per requirement)
    - department includes 'Support' => support
    Defaults to teaching if unknown.
    """
    details = load_faculty_details().get((faculty_id or '').strip().upper()) or {}
    category = (details.get('category') or '').strip().lower()
    department = (details.get('department') or '').strip().lower()
    if category == 'teaching faculty':
        return 'teaching'
    if 'non-teaching' in category:
        if 'computer applications' in department:
            return 'support'
        if 'support' in department:
            return 'support'
        if 'administrative' in department or 'administration' in department:
            return 'admin'
        # default non-teaching → admin
        return 'admin'
    # fallback
    return 'teaching'

def get_thresholds_for(date_obj: datetime, staff_type: str) -> Tuple[dt_time, dt_time]:
    weekday = date_obj.weekday()  # 0=Mon ... 5=Sat ... 6=Sun
    is_saturday = (weekday == 5)
    if staff_type == 'teaching':
        checkin_deadline = dt_time(9, 25)
        checkout_after = dt_time(13, 0) if is_saturday else dt_time(16, 30)
    elif staff_type == 'admin':
        checkin_deadline = dt_time(9, 15)
        checkout_after = dt_time(13, 30) if is_saturday else dt_time(17, 15)
    else:  # support
        checkin_deadline = dt_time(8, 30)
        checkout_after = dt_time(13, 30) if is_saturday else dt_time(17, 15)
    return checkin_deadline, checkout_after

def parse_ts(value: Any) -> datetime | None:
    if not value:
        return None
    if isinstance(value, (int, float)):
        try:
            return datetime.fromtimestamp(float(value) / (1000.0 if float(value) > 1e12 else 1.0))
        except Exception:
            return None
    if isinstance(value, str):
        for fmt in ("%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
            try:
                return datetime.strptime(value, fmt)
            except Exception:
                continue
        try:
            return datetime.fromisoformat(value)
        except Exception:
            return None
    return None

def _seconds_to_hhmmss(total_seconds: int) -> str:
    if total_seconds < 0:
        total_seconds = 0
    hours = total_seconds // 3600
    minutes_only = (total_seconds % 3600) // 60
    seconds_only = total_seconds % 60
    return f"{hours:02d}:{minutes_only:02d}:{seconds_only:02d}"


def compute_daily_delay_for_records(records: List[Dict[str, Any]], date_obj: datetime, staff_type: str) -> str:
    """Compute total delay as HH:MM:SS for a day's records for a single faculty.
    Uses earliest check-in and latest check-out. If no checkout present in any record, returns 'N/A'.
    """
    checkins: List[datetime] = []
    checkouts: List[datetime] = []
    for r in records:
        ci = parse_ts(r.get('checkin'))
        co = parse_ts(r.get('checkout'))
        if ci:
            checkins.append(ci)
        if co:
            checkouts.append(co)
    if not checkins:
        return _seconds_to_hhmmss(0)  # no checkin → treat as 0 delay
    earliest_in = min(checkins)
    checkin_deadline, checkout_after = get_thresholds_for(date_obj, staff_type)
    late_seconds = 0
    if earliest_in:
        deadline_dt = datetime.combine(earliest_in.date(), checkin_deadline)
        if earliest_in > deadline_dt:
            late_seconds = round((earliest_in - deadline_dt).total_seconds())
    if not checkouts:
        # If no checkouts, return only check-in delay (not 'N/A')
        return _seconds_to_hhmmss(late_seconds)
    latest_out = max(checkouts)
    required_dt = datetime.combine(latest_out.date(), checkout_after)
    early_leave_seconds = 0
    if latest_out < required_dt:
        early_leave_seconds = int(math.ceil((required_dt - latest_out).total_seconds()))
    return _seconds_to_hhmmss(late_seconds + early_leave_seconds)

def annotate_file_with_delay(json_file: Path) -> None:
    """Read a date file, compute delay per faculty for that date, assign same delay to all their records."""
    try:
        basename = json_file.stem  # YYYY-MM-DD
        date_obj = datetime.strptime(basename, "%Y-%m-%d")
    except Exception:
        return
    try:
        with json_file.open('r', encoding='utf-8') as f:
            data = json.load(f)
        if not isinstance(data, list):
            return
        # normalize possible 'timestamp' only schema → treat as checkin-only
        for row in data:
            if 'checkin' not in row and isinstance(row.get('timestamp'), str):
                row['checkin'] = row.get('timestamp')
            if 'checkout' not in row:
                row['checkout'] = row.get('checkout', '')
        # group by student_id
        by_id: Dict[str, List[Dict[str, Any]]] = {}
        for row in data:
            sid = (row.get('student_id') or '').strip().upper()
            by_id.setdefault(sid, []).append(row)
        # compute and annotate
        for sid, rows in by_id.items():
            staff_type = determine_staff_type(sid)
            delay_val = compute_daily_delay_for_records(rows, date_obj, staff_type)
            for r in rows:
                r['delay'] = delay_val
        with json_file.open('w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        # ignore annotation errors
        pass

def annotate_all_existing_files() -> Dict[str, Any]:
    """Annotate all JSON files currently present in source directory."""
    results = {"annotated": [], "errors": []}
    try:
        source_dir = get_attendance_dir()
        for jf in source_dir.glob('*.json'):
            try:
                annotate_file_with_delay(jf)
                results["annotated"].append(jf.name)
            except Exception as e:
                results["errors"].append({"file": jf.name, "error": str(e)})
    except Exception as e:
        results["errors"].append({"file": "*", "error": str(e)})
    return results



@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        if username == "bbhcadmin" and password == "123456":
            session["logged_in"] = True
            return redirect(url_for("index"))
        return render_template("login.html", error="Invalid username or password")
    if is_logged_in():
        return redirect(url_for("index"))
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
def index():
    if not is_logged_in():
        return redirect(url_for("login"))
    return render_template("index.html")


@app.route("/api/attendance")
def api_attendance():
    if not is_logged_in():
        return jsonify({"error": "Not authenticated"}), 401
    # Use current date by default
    date_str = request.args.get("date")
    if not date_str:
        date_str = datetime.now().strftime("%Y-%m-%d")

    # Get data directly from source directory
    rows = get_attendance_from_source(date_str)
    
    # Calculate delays for all records
    if rows:
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            # Group records by faculty ID
            by_id = {}
            for row in rows:
                sid = (row.get('student_id') or '').strip().upper()
                if sid:
                    by_id.setdefault(sid, []).append(row)
            
            # Calculate delay for each faculty member
            for sid, faculty_rows in by_id.items():
                staff_type = determine_staff_type(sid)
                delay_val = compute_daily_delay_for_records(faculty_rows, date_obj, staff_type)
                for r in faculty_rows:
                    r['delay'] = delay_val
        except Exception as e:
            print(f"Error calculating delays: {e}")

    # Get all faculty members from faculty_detail.json
    faculty_details = load_faculty_details()
    
    # Create a comprehensive list with all faculty members
    all_faculty_records = []
    
    # First, add existing attendance records with normalized case
    for row in rows:
        # Normalize student_id to uppercase for consistency
        if 'student_id' in row:
            row['student_id'] = row['student_id'].upper()
        all_faculty_records.append(row)
    
    # Then, add faculty members who don't have attendance records
    existing_faculty_ids = {row.get('student_id', '').strip().upper() for row in rows}
    
    for faculty_id, faculty_info in faculty_details.items():
        if faculty_id.upper() not in existing_faculty_ids:
            # Create a record for faculty with no attendance data
            record = {
                'student_id': faculty_id,
                'name': faculty_info.get('name', ''),
                'checkin': '',
                'checkout': '',
                'delay': 'N/A'
            }
            all_faculty_records.append(record)
    
    # Apply refined delay display logic (same as PDF)
    for record in all_faculty_records:
        # Format check-in time
        checkin = record.get('checkin', '')
        if checkin and checkin != '':
            try:
                dt = parse_ts(checkin)
                if dt:
                    checkin_formatted = dt.strftime('%H:%M:%S')
                else:
                    checkin_formatted = 'Not recorded'
            except:
                checkin_formatted = 'Not recorded'
        else:
            checkin_formatted = 'Not recorded'
        
        # Format check-out time
        checkout = record.get('checkout', '')
        if checkout and checkout != '':
            try:
                dt = parse_ts(checkout)
                if dt:
                    checkout_formatted = dt.strftime('%H:%M:%S')
                else:
                    checkout_formatted = 'Not recorded'
            except:
                checkout_formatted = 'Not recorded'
        else:
            checkout_formatted = 'Not recorded'
        
        # Apply refined delay display logic
        delay = record.get('delay', '')
        if checkin_formatted == 'Not recorded':
            # If no check-in at all, show as Absent
            record['delay'] = 'Absent'
        elif checkout_formatted == 'Not recorded':
            # If has check-in but no check-out, show as Absent
            record['delay'] = 'Absent'
        elif delay and delay != '' and delay != 'N/A':
            # If both check-in and check-out present, show calculated delay
            record['delay'] = delay
        else:
            # Fallback case - both present but no delay calculated
            record['delay'] = '00:00:00'
    
    # Sort by faculty ID
    all_faculty_records.sort(key=lambda x: (x.get('student_id') or '').strip().upper())
    
    return jsonify(all_faculty_records)




@app.route("/api/attendance/update", methods=["POST"])
def api_update_attendance():
    """Update attendance record with new check-in/check-out times."""
    if not is_logged_in():
        return jsonify({"success": False, "error": "Not authenticated"}), 401
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "No data provided"})
        
        faculty_id = data.get('faculty_id', '').strip().upper()
        date_str = data.get('date', '').strip()
        
        # Only process checkin if it exists in the request data
        new_checkin = None
        if 'checkin' in data:
            new_checkin = data.get('checkin', '').strip() if data.get('checkin') else ''
        
        # Only process checkout if it exists in the request data  
        new_checkout = None
        if 'checkout' in data:
            new_checkout = data.get('checkout', '').strip() if data.get('checkout') else ''
        
        if not faculty_id or not date_str:
            return jsonify({"success": False, "error": "Faculty ID and date are required"})
        
        # Load the attendance file directly from source
        source_dir = get_attendance_dir()
        source_file = source_dir / f"{date_str}.json"
        if not source_file.exists():
            return jsonify({"success": False, "error": f"Attendance file for {date_str} not found"})
        
        # Read current data
        with source_file.open('r', encoding='utf-8') as f:
            records = json.load(f)
        
        if not isinstance(records, list):
            return jsonify({"success": False, "error": "Invalid data format"})
        
        # Find and update the specific record
        updated = False
        record_found = False
        
        # If we're updating check-in, find the record that has the current check-in time
        # If we're updating check-out, find the record that has the current check-out time
        target_record = None
        
        if new_checkin is not None:
            # Find record with matching check-in time (or empty check-in if we're adding one)
            current_checkin_time = data.get('current_checkin', '')
            for record in records:
                if record.get('student_id', '').strip().upper() == faculty_id:
                    record_checkin = record.get('checkin', '')
                    # Convert to time-only format for comparison
                    record_checkin_time = ''
                    if record_checkin:
                        try:
                            dt = parse_ts(record_checkin)
                            if dt:
                                record_checkin_time = dt.strftime("%H:%M:%S")
                        except:
                            pass
                    
                    # Match if current check-in time matches what we're trying to edit
                    if record_checkin_time == current_checkin_time:
                        target_record = record
                        break
        elif new_checkout is not None:
            # Find record with matching check-out time (or empty check-out if we're adding one)
            current_checkout_time = data.get('current_checkout', '')
            current_checkin_time = data.get('current_checkin', '')
            
            for record in records:
                if record.get('student_id', '').strip().upper() == faculty_id:
                    record_checkout = record.get('checkout', '')
                    record_checkin = record.get('checkin', '')
                    
                    # Convert to time-only format for comparison
                    record_checkout_time = ''
                    if record_checkout:
                        try:
                            dt = parse_ts(record_checkout)
                            if dt:
                                record_checkout_time = dt.strftime("%H:%M:%S")
                        except:
                            pass
                    
                    record_checkin_time = ''
                    if record_checkin:
                        try:
                            dt = parse_ts(record_checkin)
                            if dt:
                                record_checkin_time = dt.strftime("%H:%M:%S")
                        except:
                            pass
                    
                    # Match if current check-out time matches AND check-in time matches (if provided)
                    checkout_match = record_checkout_time == current_checkout_time
                    checkin_match = True  # Default to True if no current_checkin provided
                    if current_checkin_time:
                        checkin_match = record_checkin_time == current_checkin_time
                    
                    if checkout_match and checkin_match:
                        target_record = record
                        break
        
        # If no specific record found, fall back to first match (for backward compatibility)
        if not target_record:
            for record in records:
                if record.get('student_id', '').strip().upper() == faculty_id:
                    target_record = record
                    break
        
        if target_record:
            record_found = True
            # Update check-in time if provided
            if new_checkin is not None:
                if new_checkin.strip() == '':
                    # Clear the check-in time
                    target_record['checkin'] = ''
                else:
                    # Convert HH:MM:SS format to full datetime
                    try:
                        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                        time_obj = datetime.strptime(new_checkin, "%H:%M:%S").time()
                        full_datetime = datetime.combine(date_obj, time_obj)
                        target_record['checkin'] = full_datetime.isoformat()
                    except ValueError:
                        return jsonify({"success": False, "error": "Invalid check-in time format. Use HH:MM:SS"})
            
            # Update check-out time if provided
            if new_checkout is not None:
                if new_checkout.strip() == '':
                    # Clear the check-out time
                    target_record['checkout'] = ''
                else:
                    try:
                        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                        time_obj = datetime.strptime(new_checkout, "%H:%M:%S").time()
                        full_datetime = datetime.combine(date_obj, time_obj)
                        target_record['checkout'] = full_datetime.isoformat()
                    except ValueError:
                        return jsonify({"success": False, "error": "Invalid check-out time format. Use HH:MM:SS"})
            
            updated = True
        
        # If record not found, create a new one
        if not record_found:
            # Get faculty details
            faculty_details = load_faculty_details()
            faculty_info = faculty_details.get(faculty_id.upper(), {})
            
            # Create new record
            new_record = {
                'student_id': faculty_id,
                'name': faculty_info.get('name', ''),
                'checkin': '',
                'checkout': '',
                'delay': 'N/A'
            }
            
            # Update check-in time if provided
            if new_checkin is not None:
                if new_checkin.strip() == '':
                    # Keep empty check-in time
                    new_record['checkin'] = ''
                else:
                    try:
                        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                        time_obj = datetime.strptime(new_checkin, "%H:%M:%S").time()
                        full_datetime = datetime.combine(date_obj, time_obj)
                        new_record['checkin'] = full_datetime.isoformat()
                    except ValueError:
                        return jsonify({"success": False, "error": "Invalid check-in time format. Use HH:MM:SS"})
            
            # Update check-out time if provided
            if new_checkout is not None:
                if new_checkout.strip() == '':
                    # Keep empty check-out time
                    new_record['checkout'] = ''
                else:
                    try:
                        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                        time_obj = datetime.strptime(new_checkout, "%H:%M:%S").time()
                        full_datetime = datetime.combine(date_obj, time_obj)
                        new_record['checkout'] = full_datetime.isoformat()
                    except ValueError:
                        return jsonify({"success": False, "error": "Invalid check-out time format. Use HH:MM:SS"})
            
            # Add new record to the list
            records.append(new_record)
            updated = True
        
        # Sort records by faculty ID, then by check-in time
        records.sort(key=lambda x: (
            (x.get('student_id') or '').strip().upper(),
            x.get('checkin', '') or ''
        ))
        
        # Recalculate delays for this faculty
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        faculty_records = [r for r in records if r.get('student_id', '').strip().upper() == faculty_id.upper()]
        if faculty_records:
            staff_type = determine_staff_type(faculty_id)
            delay_val = compute_daily_delay_for_records(faculty_records, date_obj, staff_type)
            # Update delay for all records of this faculty
            for record in records:
                if record.get('student_id', '').strip().upper() == faculty_id:
                    record['delay'] = delay_val
        
        # Save updated data directly to source file
        with source_file.open('w', encoding='utf-8') as f:
            json.dump(records, f, ensure_ascii=False, indent=2)
        
        return jsonify({"success": True, "message": "Attendance updated successfully"})
        
    except Exception as e:
        return jsonify({"success": False, "error": f"Failed to update attendance: {str(e)}"})


@app.route("/api/attendance/delete", methods=["POST"])
def api_delete_attendance():
    """Delete attendance record for a specific faculty member and date."""
    if not is_logged_in():
        return jsonify({"success": False, "error": "Not authenticated"}), 401
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "No data provided"})
        
        faculty_id = data.get('faculty_id', '').strip().upper()
        date_str = data.get('date', '').strip()
        
        if not faculty_id or not date_str:
            return jsonify({"success": False, "error": "Faculty ID and date are required"})
        
        # Load the attendance file directly from source
        source_dir = get_attendance_dir()
        source_file = source_dir / f"{date_str}.json"
        if not source_file.exists():
            return jsonify({"success": False, "error": f"Attendance file for {date_str} not found"})
        
        # Read current data
        with source_file.open('r', encoding='utf-8') as f:
            records = json.load(f)
        
        if not isinstance(records, list):
            return jsonify({"success": False, "error": "Invalid data format"})
        
        # Find and remove all records for this faculty member
        original_count = len(records)
        records = [r for r in records if r.get('student_id', '').strip().upper() != faculty_id]
        removed_count = original_count - len(records)
        
        if removed_count == 0:
            return jsonify({"success": False, "error": f"No records found for faculty {faculty_id} on {date_str}"})
        
        # Create a placeholder record for the deleted faculty member
        faculty_details = load_faculty_details()
        faculty_info = faculty_details.get(faculty_id.upper(), {})
        
        placeholder_record = {
            'student_id': faculty_id,
            'name': faculty_info.get('name', ''),
            'checkin': '',
            'checkout': '',
            'delay': 'N/A'
        }
        
        # Insert the placeholder record in the correct sorted position
        records.append(placeholder_record)
        records.sort(key=lambda x: (x.get('student_id') or '').strip().upper())
        
        # Save updated data directly to source file
        with source_file.open('w', encoding='utf-8') as f:
            json.dump(records, f, ensure_ascii=False, indent=2)
        
        return jsonify({
            "success": True, 
            "message": f"Successfully deleted {removed_count} record(s) for {faculty_id} and created placeholder record",
            "deleted_count": removed_count,
            "placeholder_created": True
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": f"Failed to delete attendance: {str(e)}"})


@app.route("/api/faculty-list")
def api_faculty_list():
    """Get list of all faculty members for dropdown."""
    if not is_logged_in():
        return jsonify({"error": "Not authenticated"}), 401
    
    try:
        faculty_details = load_faculty_details()
        faculty_list = []
        
        for faculty_id, faculty_info in faculty_details.items():
            faculty_list.append({
                "id": faculty_id,
                "name": faculty_info.get("name", "")
            })
        
        # Sort by faculty ID
        faculty_list.sort(key=lambda x: x["id"])
        
        return jsonify(faculty_list)
    except Exception as e:
        return jsonify({"error": f"Failed to load faculty list: {str(e)}"}), 500


@app.route("/api/attendance/add", methods=["POST"])
def api_add_attendance():
    """Add a new attendance record for a faculty member."""
    if not is_logged_in():
        return jsonify({"success": False, "error": "Not authenticated"}), 401
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "No data provided"})
        
        faculty_id = data.get('faculty_id', '').strip().upper()
        date_str = data.get('date', '').strip()
        checkin_time = data.get('checkin', '').strip()
        checkout_time = data.get('checkout', '').strip()
        
        if not faculty_id or not date_str:
            return jsonify({"success": False, "error": "Faculty ID and date are required"})
        
        # Load the attendance file directly from source
        source_dir = get_attendance_dir()
        source_file = source_dir / f"{date_str}.json"
        
        # Read current data
        records = []
        if source_file.exists():
            with source_file.open('r', encoding='utf-8') as f:
                records = json.load(f)
        
        if not isinstance(records, list):
            records = []
        
        # Get faculty details
        faculty_details = load_faculty_details()
        faculty_info = faculty_details.get(faculty_id.upper(), {})
        
        # Create new record with provided times
        new_record = {
            'student_id': faculty_id,
            'name': faculty_info.get('name', ''),
            'checkin': '',
            'checkout': '',
            'delay': 'N/A'
        }
        
        # Set check-in time if provided
        if checkin_time:
            try:
                date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                time_obj = datetime.strptime(checkin_time, "%H:%M:%S").time()
                full_datetime = datetime.combine(date_obj, time_obj)
                new_record['checkin'] = full_datetime.isoformat()
            except ValueError:
                return jsonify({"success": False, "error": "Invalid check-in time format. Use HH:MM:SS"})
        
        # Set check-out time if provided
        if checkout_time:
            try:
                date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                time_obj = datetime.strptime(checkout_time, "%H:%M:%S").time()
                full_datetime = datetime.combine(date_obj, time_obj)
                new_record['checkout'] = full_datetime.isoformat()
            except ValueError:
                return jsonify({"success": False, "error": "Invalid check-out time format. Use HH:MM:SS"})
        
        # Add new record
        records.append(new_record)
        
        # Sort all records by faculty ID, then by check-in time
        records.sort(key=lambda x: (
            (x.get('student_id') or '').strip().upper(),
            x.get('checkin', '') or ''
        ))
        
        # Recalculate delays for this faculty
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        faculty_records = [r for r in records if r.get('student_id', '').strip().upper() == faculty_id.upper()]
        if faculty_records:
            staff_type = determine_staff_type(faculty_id)
            delay_val = compute_daily_delay_for_records(faculty_records, date_obj, staff_type)
            # Update delay for all records of this faculty
            for record in records:
                if record.get('student_id', '').strip().upper() == faculty_id:
                    record['delay'] = delay_val
        
        # Save updated data directly to source file
        with source_file.open('w', encoding='utf-8') as f:
            json.dump(records, f, ensure_ascii=False, indent=2)
        
        return jsonify({
            "success": True, 
            "message": f"Successfully added new record for {faculty_id}"
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": f"Failed to add attendance: {str(e)}"})


@app.route("/api/attendance/delete-specific", methods=["POST"])
def api_delete_specific_attendance():
    """Delete a specific attendance record by index."""
    if not is_logged_in():
        return jsonify({"success": False, "error": "Not authenticated"}), 401
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "error": "No data provided"})
        
        faculty_id = data.get('faculty_id', '').strip().upper()
        date_str = data.get('date', '').strip()
        record_index = data.get('record_index', 0)
        
        if not faculty_id or not date_str:
            return jsonify({"success": False, "error": "Faculty ID and date are required"})
        
        # Load the attendance file directly from source
        source_dir = get_attendance_dir()
        source_file = source_dir / f"{date_str}.json"
        if not source_file.exists():
            return jsonify({"success": False, "error": f"Attendance file for {date_str} not found"})
        
        # Read current data
        with source_file.open('r', encoding='utf-8') as f:
            records = json.load(f)
        
        if not isinstance(records, list):
            return jsonify({"success": False, "error": "Invalid data format"})
        
        # Find all records for this faculty member
        faculty_records = [i for i, r in enumerate(records) if r.get('student_id', '').strip().upper() == faculty_id.upper()]
        
        if not faculty_records:
            return jsonify({"success": False, "error": f"No records found for faculty {faculty_id}"})
        
        if record_index >= len(faculty_records):
            return jsonify({"success": False, "error": "Invalid record index"})
        
        # Get the actual record index in the full list
        actual_index = faculty_records[record_index]
        
        # Remove the specific record
        deleted_record = records.pop(actual_index)
        
        # If no more records exist for this faculty, create a placeholder
        remaining_faculty_records = [r for r in records if r.get('student_id', '').strip().upper() == faculty_id.upper()]
        
        if not remaining_faculty_records:
            # Create placeholder record
            faculty_details = load_faculty_details()
            faculty_info = faculty_details.get(faculty_id.upper(), {})
            
            placeholder_record = {
                'student_id': faculty_id,
                'name': faculty_info.get('name', ''),
                'checkin': '',
                'checkout': '',
                'delay': 'N/A'
            }
            
            records.append(placeholder_record)
            placeholder_created = True
        else:
            placeholder_created = False
        
        # Sort all records by faculty ID
        records.sort(key=lambda x: (x.get('student_id') or '').strip().upper())
        
        # Save updated data directly to source file
        with source_file.open('w', encoding='utf-8') as f:
            json.dump(records, f, ensure_ascii=False, indent=2)
        
        return jsonify({
            "success": True, 
            "message": f"Successfully deleted record for {faculty_id}",
            "placeholder_created": placeholder_created
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": f"Failed to delete attendance: {str(e)}"})


@app.route("/api/monthly-delay-report", methods=["GET"])
def api_monthly_delay_report():
    """Get monthly delay report data."""
    if not is_logged_in():
        return jsonify({"success": False, "error": "Not authenticated"}), 401
    
    try:
        month = request.args.get('month')
        year = request.args.get('year')
        
        if not month or not year:
            return jsonify({"success": False, "error": "Month and year are required"})
        
        # Get monthly delay data
        monthly_delays = []
        faculty_details = load_faculty_details()
        
        for faculty_id, faculty_info in faculty_details.items():
            total_delay_seconds = 0
            days_with_records = 0
            
            for day in range(1, 32):
                try:
                    date_str = f"{year}-{month.zfill(2)}-{day:02d}"
                    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                    
                    source_dir = get_attendance_dir()
                    source_file = source_dir / f"{date_str}.json"
                    
                    if source_file.exists():
                        with source_file.open('r', encoding='utf-8') as f:
                            records = json.load(f)
                        
                        faculty_records = [r for r in records if r.get('student_id', '').strip().upper() == faculty_id.upper()]
                        
                        if faculty_records:
                            days_with_records += 1
                            staff_type = determine_staff_type(faculty_id)
                            daily_delay = compute_daily_delay_for_records(faculty_records, date_obj, staff_type)
                            
                            if daily_delay != 'N/A':
                                delay_parts = daily_delay.split(':')
                                if len(delay_parts) == 3:
                                    hours, minutes, seconds = map(int, delay_parts)
                                    total_delay_seconds += hours * 3600 + minutes * 60 + seconds
                
                except ValueError:
                    continue
            
            total_delay = 'N/A' if days_with_records == 0 else _seconds_to_hhmmss(total_delay_seconds)
            
            monthly_delays.append({
                'faculty_id': faculty_id,
                'name': faculty_info.get('name', ''),
                'total_delay': total_delay
            })
        
        monthly_delays.sort(key=lambda x: x['faculty_id'])
        
        return jsonify({
            "success": True,
            "data": monthly_delays,
            "month": month,
            "year": year
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": f"Failed to get monthly delay report: {str(e)}"})


@app.route("/api/monthly-delay-report/excel", methods=["GET"])
def api_monthly_delay_report_excel():
    """Export monthly delay report as Excel file."""
    if not is_logged_in():
        return jsonify({"success": False, "error": "Not authenticated"}), 401
    
    try:
        month = request.args.get('month')
        year = request.args.get('year')
        
        if not month or not year:
            return jsonify({"success": False, "error": "Month and year are required"})
        
        # Get monthly delay data
        monthly_delays = []
        faculty_details = load_faculty_details()
        
        for faculty_id, faculty_info in faculty_details.items():
            total_delay_seconds = 0
            days_with_records = 0
            
            for day in range(1, 32):
                try:
                    date_str = f"{year}-{month.zfill(2)}-{day:02d}"
                    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                    
                    source_dir = get_attendance_dir()
                    source_file = source_dir / f"{date_str}.json"
                    
                    if source_file.exists():
                        with source_file.open('r', encoding='utf-8') as f:
                            records = json.load(f)
                        
                        faculty_records = [r for r in records if r.get('student_id', '').strip().upper() == faculty_id.upper()]
                        
                        if faculty_records:
                            days_with_records += 1
                            staff_type = determine_staff_type(faculty_id)
                            daily_delay = compute_daily_delay_for_records(faculty_records, date_obj, staff_type)
                            
                            if daily_delay != 'N/A':
                                delay_parts = daily_delay.split(':')
                                if len(delay_parts) == 3:
                                    hours, minutes, seconds = map(int, delay_parts)
                                    total_delay_seconds += hours * 3600 + minutes * 60 + seconds
                
                except ValueError:
                    continue
            
            total_delay = 'N/A' if days_with_records == 0 else _seconds_to_hhmmss(total_delay_seconds)
            
            monthly_delays.append({
                'faculty_id': faculty_id,
                'name': faculty_info.get('name', ''),
                'total_delay': total_delay
            })
        
        monthly_delays.sort(key=lambda x: x['faculty_id'])
        
        # Create Excel workbook
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Monthly Delay Report {month}-{year}"
        
        # Convert month number to month name
        month_names = {
            '01': 'January', '02': 'February', '03': 'March', '04': 'April',
            '05': 'May', '06': 'June', '07': 'July', '08': 'August',
            '09': 'September', '10': 'October', '11': 'November', '12': 'December'
        }
        month_name = month_names.get(month.zfill(2), month)
        
        # Meta table - Report Information
        meta_data = [
            ['Report Type', 'Monthly Delay Report'],
            ['Month', month_name],
            ['Year', year],
            ['Generated On', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Total Faculty', len(monthly_delays)],
            ['', ''],  # Empty row for spacing
        ]
        
        # Add meta table
        for row_idx, (key, value) in enumerate(meta_data, 1):
            if key and value:  # Skip empty rows
                # Key column
                key_cell = ws.cell(row=row_idx, column=1, value=key)
                key_cell.font = Font(bold=True, size=11)
                key_cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                key_cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                
                # Value column
                value_cell = ws.cell(row=row_idx, column=2, value=value)
                value_cell.font = Font(size=11)
                value_cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
                value_cell.alignment = Alignment(horizontal="left", vertical="center")
                value_cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
        
        # Main data table starts after meta table
        data_start_row = len([item for item in meta_data if item[0] and item[1]]) + 3
        
        # Main table headers
        headers = ['Faculty ID', 'Name', 'Total Delay']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=data_start_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='medium', color='000000'),
                right=Side(style='medium', color='000000'),
                top=Side(style='medium', color='000000'),
                bottom=Side(style='medium', color='000000')
            )
        
        # Data rows
        for row, delay_data in enumerate(monthly_delays, data_start_row + 1):
            # Faculty ID
            id_cell = ws.cell(row=row, column=1, value=delay_data['faculty_id'])
            id_cell.font = Font(bold=True, size=11)
            id_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Name
            name_cell = ws.cell(row=row, column=2, value=delay_data['name'])
            name_cell.font = Font(size=11)
            name_cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Total Delay
            delay_cell = ws.cell(row=row, column=3, value=delay_data['total_delay'])
            delay_cell.font = Font(size=11)
            delay_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add borders and alternating row colors
            row_color = "FFFFFF" if row % 2 == 0 else "F8F9FA"
            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return app.response_class(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=monthly_delay_report_{year}_{month}.xlsx'}
        )
        
    except Exception as e:
        return jsonify({"success": False, "error": f"Failed to export Excel: {str(e)}"})


@app.route("/api/monthly-delay-report/pdf", methods=["GET"])
def api_monthly_delay_report_pdf():
    """Export monthly delay report as PDF file."""
    if not is_logged_in():
        return jsonify({"success": False, "error": "Not authenticated"}), 401
    
    try:
        month = request.args.get('month')
        year = request.args.get('year')
        
        if not month or not year:
            return jsonify({"success": False, "error": "Month and year are required"})
        
        # Get monthly delay data
        monthly_delays = []
        faculty_details = load_faculty_details()
        
        for faculty_id, faculty_info in faculty_details.items():
            total_delay_seconds = 0
            days_with_records = 0
            
            for day in range(1, 32):
                try:
                    date_str = f"{year}-{month.zfill(2)}-{day:02d}"
                    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                    
                    source_dir = get_attendance_dir()
                    source_file = source_dir / f"{date_str}.json"
                    
                    if source_file.exists():
                        with source_file.open('r', encoding='utf-8') as f:
                            records = json.load(f)
                        
                        faculty_records = [r for r in records if r.get('student_id', '').strip().upper() == faculty_id.upper()]
                        
                        if faculty_records:
                            days_with_records += 1
                            staff_type = determine_staff_type(faculty_id)
                            daily_delay = compute_daily_delay_for_records(faculty_records, date_obj, staff_type)
                            
                            if daily_delay != 'N/A':
                                delay_parts = daily_delay.split(':')
                                if len(delay_parts) == 3:
                                    hours, minutes, seconds = map(int, delay_parts)
                                    total_delay_seconds += hours * 3600 + minutes * 60 + seconds
                
                except ValueError:
                    continue
            
            total_delay = 'N/A' if days_with_records == 0 else _seconds_to_hhmmss(total_delay_seconds)
            
            monthly_delays.append({
                'faculty_id': faculty_id,
                'name': faculty_info.get('name', ''),
                'total_delay': total_delay
            })
        
        monthly_delays.sort(key=lambda x: x['faculty_id'])
        
        # Create PDF with same template as daily attendance export
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from io import BytesIO
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Build content
        story = []
        
        # Use the new header design
        header_table = create_pdf_header()
        story.append(header_table)
        story.append(Spacer(1, 20))  # Space after header
        
        # Convert month number to month name
        month_names = {
            '01': 'January', '02': 'February', '03': 'March', '04': 'April',
            '05': 'May', '06': 'June', '07': 'July', '08': 'August',
            '09': 'September', '10': 'October', '11': 'November', '12': 'December'
        }
        month_name = month_names.get(month.zfill(2), month)
        
        # Report title (centered with proper spacing)
        report_title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading2'],
            fontSize=16,  # Smaller font to save space
            spaceAfter=12,  # Restore proper spacing
            alignment=1,  # Center alignment
            fontName='Helvetica-Bold',
            textColor=colors.black
        )
        report_title = Paragraph("Monthly Delay Report", report_title_style)
        story.append(report_title)
        
        # Month and year (centered with proper spacing)
        month_style = ParagraphStyle(
            'MonthStyle',
            parent=styles['Normal'],
            fontSize=14,  # Smaller font to save space
            spaceAfter=20,  # Restore proper spacing
            alignment=1,  # Center alignment
            textColor=colors.grey,  # Gray color for subtitle
            fontName='Helvetica'
        )
        month_text = Paragraph(f"Month: {month_name} {year}", month_style)
        story.append(month_text)
        story.append(Spacer(1, 20))  # Restore proper spacing
        
        # Create table data
        table_data = [['Faculty ID', 'Name', 'Total Delay']]
        for delay_data in monthly_delays:
            table_data.append([
                delay_data['faculty_id'],
                delay_data['name'],
                delay_data['total_delay']
            ])
        
        # Create table with larger fonts and professional styling
        table = Table(table_data, colWidths=[90, 350, 90])  # Better column widths
        table.setStyle(TableStyle([
            # Header styling - black background with white text
            ('BACKGROUND', (0, 0), (-1, 0), colors.black),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'LEFT'),  # Left align headers
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),  # Larger header font
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),  # Better padding
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            # White vertical borders for header
            ('LINEBEFORE', (0, 0), (0, 0), 1.0, colors.white),
            ('LINEBEFORE', (1, 0), (1, 0), 1.0, colors.white),
            ('LINEBEFORE', (2, 0), (2, 0), 1.0, colors.white),
            ('LINEAFTER', (0, 0), (0, 0), 1.0, colors.white),
            ('LINEAFTER', (1, 0), (1, 0), 1.0, colors.white),
            ('LINEAFTER', (2, 0), (2, 0), 1.0, colors.white),
            # Data row styling - alternating white and light grey
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),  # Left align data
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 11),  # Larger data font for better readability
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),  # Better padding
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            # Black borders for all cells
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Clean borders
        ]))
        
        story.append(table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        return app.response_class(
            buffer.getvalue(),
            mimetype='application/pdf',
            headers={'Content-Disposition': f'attachment; filename=monthly_delay_report_{year}_{month}.pdf'}
        )
        
    except Exception as e:
        return jsonify({"success": False, "error": f"Failed to export PDF: {str(e)}"})


@app.route("/api/daily-attendance-report/excel", methods=["GET"])
def api_daily_attendance_report_excel():
    """Export daily attendance report as Excel file."""
    if not is_logged_in():
        return jsonify({"success": False, "error": "Not authenticated"}), 401
    
    try:
        date = request.args.get('date')
        if not date:
            return jsonify({"success": False, "error": "Date is required"})
        
        # Get attendance data for the date
        source_dir = get_attendance_dir()
        source_file = source_dir / f"{date}.json"
        
        if not source_file.exists():
            return jsonify({"success": False, "error": f"No attendance data found for {date}"})
        
        with source_file.open('r', encoding='utf-8') as f:
            records = json.load(f)
        
        # Sort records by faculty ID
        records.sort(key=lambda x: (x.get('student_id') or '').strip().upper())
        
        # Create Excel workbook with ONLY 5 columns using a different approach
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Faculty Attendance Report {date}"
        
        # CRITICAL: Only work with columns A-E, ignore everything else
        # Set column dimensions to 0 for all columns beyond E
        for col in range(6, 50):  # Set columns F onwards to 0 width
            ws.column_dimensions[chr(64 + col)].width = 0
            ws.column_dimensions[chr(64 + col)].hidden = True
        
        # Convert date format from YYYY-MM-DD to DD-MM-YYYY
        date_parts = date.split('-')
        formatted_date = f"{date_parts[2]}-{date_parts[1]}-{date_parts[0]}"
        
        # Report title
        title_cell = ws.cell(row=1, column=1, value=f"Faculty Attendance Report - {formatted_date}")
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add some spacing
        ws.row_dimensions[2].height = 20
        
        # Table headers (only 5 columns: Faculty ID, Name, Check-in, Check-out, Delay)
        headers = ['Faculty ID', 'Name', 'Check-in (Time)', 'Check-out (Time)', 'Delay (Time)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=12)  # White text
            cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Black background
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='medium', color='000000'),
                right=Side(style='medium', color='000000'),
                top=Side(style='medium', color='000000'),
                bottom=Side(style='medium', color='000000')
            )
        
        # Ensure any other columns beyond the 5 main columns have white background
        for col in range(6, 50):  # Columns F onwards
            cell = ws.cell(row=3, column=col, value="")
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White background
            cell.font = Font(color="000000", size=12)  # Black text
            cell.border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
        
        # Data rows (only 5 columns)
        for row, record in enumerate(records, 4):
            # Faculty ID
            id_cell = ws.cell(row=row, column=1, value=record.get('student_id', ''))
            id_cell.font = Font(bold=True, size=11)
            id_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Name
            name_cell = ws.cell(row=row, column=2, value=record.get('name', ''))
            name_cell.font = Font(size=11)
            name_cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Check-in time
            checkin = record.get('checkin', '')
            if checkin and checkin != '':
                try:
                    from datetime import datetime
                    dt = datetime.fromisoformat(checkin.replace('Z', '+00:00'))
                    checkin_formatted = dt.strftime('%H:%M:%S')
                except:
                    checkin_formatted = 'Not recorded'
            else:
                checkin_formatted = 'Not recorded'
            
            checkin_cell = ws.cell(row=row, column=3, value=checkin_formatted)
            checkin_cell.font = Font(size=11)
            checkin_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Check-out time
            checkout = record.get('checkout', '')
            if checkout and checkout != '':
                try:
                    from datetime import datetime
                    dt = datetime.fromisoformat(checkout.replace('Z', '+00:00'))
                    checkout_formatted = dt.strftime('%H:%M:%S')
                except:
                    checkout_formatted = 'Not recorded'
            else:
                checkout_formatted = 'Not recorded'
            
            checkout_cell = ws.cell(row=row, column=4, value=checkout_formatted)
            checkout_cell.font = Font(size=11)
            checkout_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Delay time
            delay = record.get('delay', '')
            if delay and delay != '' and delay != 'N/A':
                delay_formatted = delay
            elif checkin_formatted == 'Not recorded' or checkout_formatted == 'Not recorded':
                delay_formatted = 'Absent'
            else:
                delay_formatted = '00:00:00'
            
            delay_cell = ws.cell(row=row, column=5, value=delay_formatted)
            delay_cell.font = Font(size=11)
            delay_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add borders only (no background colors)
            for col in range(1, 6):  # Only 5 columns
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
        
        # Set column widths (only for 5 columns)
        column_widths = [15, 30, 15, 15, 15]  # Faculty ID, Name, Check-in, Check-out, Delay
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # Explicitly clear any styling from extra columns
        for row in range(1, len(records) + 10):
            for col in range(6, 50):  # Clear columns F onwards
                cell = ws.cell(row=row, column=col)
                cell.value = None
                cell.fill = PatternFill()  # No fill
                cell.font = Font()  # Default font
                cell.border = Border()  # No border
                cell.alignment = Alignment()  # Default alignment
        
        # Set the used range to only include columns A-E
        ws.max_column = 5
        
        # Save to BytesIO
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return app.response_class(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=daily_attendance_report_{date}.xlsx'}
        )
        
    except Exception as e:
        return jsonify({"success": False, "error": f"Failed to export Excel: {str(e)}"})

@app.route("/api/daily-attendance-report/pdf", methods=["GET"])
def api_daily_attendance_report_pdf():
    """Export daily attendance report as PDF file."""
    if not is_logged_in():
        return jsonify({"success": False, "error": "Not authenticated"}), 401
    
    try:
        date = request.args.get('date')
        if not date:
            return jsonify({"success": False, "error": "Date is required"})
        
        print(f"PDF generation requested for date: {date}")
        
        # Get attendance data for the date
        source_dir = get_attendance_dir()
        source_file = source_dir / f"{date}.json"
        
        if not source_file.exists():
            print(f"File not found: {source_file}")
            return jsonify({"success": False, "error": f"No attendance data found for {date}"})
        
        print(f"Reading data from: {source_file}")
        with source_file.open('r', encoding='utf-8') as f:
            records = json.load(f)
        
        print(f"Found {len(records)} records")
        
        # Calculate delays for all records (same logic as API)
        if records:
            try:
                date_obj = datetime.strptime(date, "%Y-%m-%d")
                # Group records by faculty ID
                by_id = {}
                for row in records:
                    sid = (row.get('student_id') or '').strip().upper()
                    if sid:
                        by_id.setdefault(sid, []).append(row)
                
                # Calculate delay for each faculty member
                for sid, faculty_rows in by_id.items():
                    staff_type = determine_staff_type(sid)
                    delay_val = compute_daily_delay_for_records(faculty_rows, date_obj, staff_type)
                    for r in faculty_rows:
                        r['delay'] = delay_val
            except Exception as e:
                print(f"Error calculating delays: {e}")
        
        # Consolidate records by faculty ID (one row per faculty)
        consolidated_records = {}
        for record in records:
            faculty_id = (record.get('student_id') or '').strip().upper()
            if faculty_id not in consolidated_records:
                consolidated_records[faculty_id] = {
                    'student_id': faculty_id,
                    'name': record.get('name', ''),
                    'checkins': [],
                    'checkouts': [],
                    'delay': record.get('delay', '')
                }
            
            # Collect all check-ins and check-outs
            checkin = record.get('checkin', '')
            checkout = record.get('checkout', '')
            
            if checkin and checkin != '':
                try:
                    dt = datetime.fromisoformat(checkin.replace('Z', '+00:00'))
                    consolidated_records[faculty_id]['checkins'].append(dt)
                except:
                    pass
            
            if checkout and checkout != '':
                try:
                    dt = datetime.fromisoformat(checkout.replace('Z', '+00:00'))
                    consolidated_records[faculty_id]['checkouts'].append(dt)
                except:
                    pass
        
        # Sort consolidated records by faculty ID
        sorted_faculty_ids = sorted(consolidated_records.keys())
        
        # Create PDF with same template as monthly report
        print("Starting PDF generation...")
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from io import BytesIO
        
        print("Creating PDF buffer...")
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Build content
        story = []
        
        # Use the new header design
        header_table = create_pdf_header()
        story.append(header_table)
        story.append(Spacer(1, 20))  # Space after header
        
        # Convert date format from YYYY-MM-DD to DD-MM-YYYY
        date_parts = date.split('-')
        formatted_date = f"{date_parts[2]}-{date_parts[1]}-{date_parts[0]}"
        
        # Report title - Two line format
        report_title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            alignment=1  # Center alignment
        )
        report_title = Paragraph("Faculty Attendance", report_title_style)
        story.append(report_title)
        
        # Date line
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Heading3'],
            fontSize=12,
            spaceAfter=20,
            alignment=1  # Center alignment
        )
        date_title = Paragraph(f"Date: {formatted_date}", date_style)
        story.append(date_title)
        story.append(Spacer(1, 20))
        
        # Create table data with consolidated records
        table_data = [['Faculty ID', 'Name', 'Check In', 'Check Out', 'Delay']]
        for faculty_id in sorted_faculty_ids:
            record = consolidated_records[faculty_id]
            
            # Format check in time (earliest check-in)
            if record['checkins']:
                earliest_checkin = min(record['checkins'])
                checkin_formatted = earliest_checkin.strftime('%H:%M:%S')
            else:
                checkin_formatted = 'Absent'
            
            # Format check out time (latest check-out)
            if record['checkouts']:
                latest_checkout = max(record['checkouts'])
                checkout_formatted = latest_checkout.strftime('%H:%M:%S')
            else:
                checkout_formatted = 'Absent'
            
            # Format delay value with refined logic for check-in without check-out
            delay = record.get('delay', '')
            if checkin_formatted == 'Absent':
                # If no check-in at all, show as Absent
                delay_formatted = 'Absent'
            elif checkout_formatted == 'Absent':
                # If has check-in but no check-out, show as Absent
                delay_formatted = 'Absent'
            elif delay and delay != '' and delay != 'N/A':
                # If both check-in and check-out present, show calculated delay
                delay_formatted = delay
            else:
                # Fallback case - both present but no delay calculated
                delay_formatted = '00:00:00'
            
            table_data.append([
                record.get('student_id', ''),
                record.get('name', ''),
                checkin_formatted,
                checkout_formatted,
                delay_formatted
            ])
        
        # Create table with professional styling
        table = Table(table_data, colWidths=[80, 150, 80, 80, 80])
        table.setStyle(TableStyle([
            # Header styling - black background with white text (matching monthly report)
            ('BACKGROUND', (0, 0), (-1, 0), colors.black),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'LEFT'),  # Left align headers (matching monthly report)
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),  # Better padding
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            # White vertical borders for header (matching monthly report)
            ('LINEBEFORE', (0, 0), (0, 0), 1.0, colors.white),
            ('LINEBEFORE', (1, 0), (1, 0), 1.0, colors.white),
            ('LINEBEFORE', (2, 0), (2, 0), 1.0, colors.white),
            ('LINEBEFORE', (3, 0), (3, 0), 1.0, colors.white),
            ('LINEBEFORE', (4, 0), (4, 0), 1.0, colors.white),
            ('LINEAFTER', (0, 0), (0, 0), 1.0, colors.white),
            ('LINEAFTER', (1, 0), (1, 0), 1.0, colors.white),
            ('LINEAFTER', (2, 0), (2, 0), 1.0, colors.white),
            ('LINEAFTER', (3, 0), (3, 0), 1.0, colors.white),
            ('LINEAFTER', (4, 0), (4, 0), 1.0, colors.white),
            # Data row styling - alternating white and light grey (matching monthly report)
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),  # Left align data (matching monthly report)
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 11),  # Larger data font for better readability
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),  # Better padding
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            # Black borders for all cells (matching monthly report)
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),  # Clean borders
        ]))
        
        story.append(table)
        
        # Build PDF
        print("Building PDF...")
        try:
            doc.build(story)
            buffer.seek(0)
            pdf_content = buffer.getvalue()
            print(f"PDF generated successfully, size: {len(pdf_content)} bytes")
        except Exception as e:
            print(f"Error building PDF: {e}")
            raise e
        
        return app.response_class(
            pdf_content,
            mimetype='application/pdf',
            headers={'Content-Disposition': f'attachment; filename=daily_attendance_report_{date}.pdf'}
        )
        
    except Exception as e:
        return jsonify({"success": False, "error": f"Failed to export PDF: {str(e)}"})


@app.route("/faculty-detailed-report")
def faculty_detailed_report_page():
    """Serve the faculty detailed report webpage."""
    if not is_logged_in():
        return redirect(url_for("login"))
    return render_template("faculty_detailed_report.html")


@app.route("/api/faculty-detailed-report", methods=["GET"])
def api_faculty_detailed_report():
    """Get faculty detailed report data."""
    if not is_logged_in():
        return jsonify({"success": False, "error": "Not authenticated"}), 401
    
    try:
        month = request.args.get('month')
        year = request.args.get('year')
        faculty = request.args.get('faculty')
        
        if not month or not year or not faculty:
            return jsonify({"success": False, "error": "Month, year, and faculty are required"})
        
        # Get faculty details
        faculty_details = load_faculty_details()
        
        # Determine which faculty to process
        if faculty == 'all':
            target_faculty = list(faculty_details.keys())
        else:
            target_faculty = [faculty]
        
        # Get all days in the month
        from calendar import monthrange
        days_in_month = monthrange(int(year), int(month))[1]
        
        # Collect data for each faculty member
        faculty_reports = []
        
        for faculty_id in target_faculty:
            faculty_info = faculty_details.get(faculty_id, {})
            faculty_name = faculty_info.get('name', '')
            
            # Collect daily data for this faculty
            daily_data = []
            total_delay_seconds = 0
            
            for day in range(1, days_in_month + 1):
                try:
                    date_str = f"{year}-{month.zfill(2)}-{day:02d}"
                    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                    
                    source_dir = get_attendance_dir()
                    source_file = source_dir / f"{date_str}.json"
                    
                    day_data = {
                        'date': date_str,
                        'checkin': 'Not recorded',
                        'checkout': 'Not recorded',
                        'delay': '00:00:00'
                    }
                    
                    if source_file.exists():
                        with source_file.open('r', encoding='utf-8') as f:
                            records = json.load(f)
                        
                        faculty_records = [r for r in records if r.get('student_id', '').strip().upper() == faculty_id.upper()]
                        
                        if faculty_records:
                            # Find earliest check-in and latest check-out
                            checkins = []
                            checkouts = []
                            
                            for record in faculty_records:
                                ci = parse_ts(record.get('checkin'))
                                co = parse_ts(record.get('checkout'))
                                if ci:
                                    checkins.append(ci)
                                if co:
                                    checkouts.append(co)
                            
                            if checkins:
                                earliest_checkin = min(checkins)
                                day_data['checkin'] = earliest_checkin.strftime('%H:%M:%S')
                            
                            if checkouts:
                                latest_checkout = max(checkouts)
                                day_data['checkout'] = latest_checkout.strftime('%H:%M:%S')
                            
                            # Calculate delay
                            staff_type = determine_staff_type(faculty_id)
                            delay_val = compute_daily_delay_for_records(faculty_records, date_obj, staff_type)
                            
                            # Check if both check-in and check-out are not recorded
                            if not checkins and not checkouts:
                                day_data['delay'] = 'Absent'
                            else:
                                day_data['delay'] = delay_val
                            
                            # Add to total delay
                            if delay_val and delay_val != 'N/A':
                                try:
                                    delay_parts = delay_val.split(':')
                                    if len(delay_parts) == 3:
                                        hours, minutes, seconds = map(int, delay_parts)
                                        total_delay_seconds += hours * 3600 + minutes * 60 + seconds
                                except:
                                    pass
                    
                    daily_data.append(day_data)
                    
                except ValueError:
                    # Invalid date, skip
                    daily_data.append({
                        'date': f"{year}-{month.zfill(2)}-{day:02d}",
                        'checkin': 'Invalid Date',
                        'checkout': 'Invalid Date',
                        'delay': '00:00:00'
                    })
            
            faculty_reports.append({
                'faculty_id': faculty_id,
                'faculty_name': faculty_name,
                'daily_data': daily_data,
                'total_delay': _seconds_to_hhmmss(total_delay_seconds)
            })
        
        return jsonify({
            "success": True,
            "data": faculty_reports,
            "month": month,
            "year": year,
            "faculty": faculty
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": f"Failed to get faculty detailed report: {str(e)}"})


@app.route("/api/faculty-detailed-report/excel", methods=["GET"])
def api_faculty_detailed_report_excel():
    """Export faculty detailed report as Excel file."""
    if not is_logged_in():
        return jsonify({"success": False, "error": "Not authenticated"}), 401
    
    try:
        month = request.args.get('month')
        year = request.args.get('year')
        faculty = request.args.get('faculty')
        
        if not month or not year or not faculty:
            return jsonify({"success": False, "error": "Month, year, and faculty are required"})
        
        # Get faculty details
        faculty_details = load_faculty_details()
        
        # Determine which faculty to process
        if faculty == 'all':
            target_faculty = list(faculty_details.keys())
        else:
            target_faculty = [faculty]
        
        # Get all days in the month
        from calendar import monthrange
        days_in_month = monthrange(int(year), int(month))[1]
        
        # Create Excel workbook
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Faculty Detailed Report {month}-{year}"
        
        # Convert month number to month name
        month_names = {
            '01': 'January', '02': 'February', '03': 'March', '04': 'April',
            '05': 'May', '06': 'June', '07': 'July', '08': 'August',
            '09': 'September', '10': 'October', '11': 'November', '12': 'December'
        }
        month_name = month_names.get(month.zfill(2), month)
        
        # Add top heading for the entire report
        top_heading = ws.cell(row=1, column=1, value=f"DETAILED FACULTY REPORT - {month_name}")
        top_heading.font = Font(bold=True, size=16, color="000000")
        top_heading.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f'A1:D1')  # Merge across all columns
        ws.row_dimensions[1].height = 25
        
        # Process each faculty member
        current_row = 2
        
        for faculty_id in target_faculty:
            faculty_info = faculty_details.get(faculty_id, {})
            faculty_name = faculty_info.get('name', '')
            
            # Faculty header - Blue text format like second image
            faculty_header = ws.cell(row=current_row, column=1, value=f"{faculty_id} - {faculty_name}")
            faculty_header.font = Font(bold=True, size=14, color="0000FF")  # Blue text
            faculty_header.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1
            
            # Create date-wise table for this faculty
            # Headers: Date, Check In Time, Check Out Time, Delay
            headers = ['Date', 'Check In Time', 'Check Out Time', 'Delay']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Add white borders for better column identification
                cell.border = Border(
                    left=Side(style='medium', color='FFFFFF'),  # White left border
                    right=Side(style='medium', color='FFFFFF'),  # White right border
                    top=Side(style='medium', color='000000'),
                    bottom=Side(style='medium', color='000000')
                )
            current_row += 1
            
            # Collect daily data for this faculty
            total_delay_seconds = 0
            
            for day in range(1, days_in_month + 1):
                try:
                    date_str = f"{year}-{month.zfill(2)}-{day:02d}"
                    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                    
                    source_dir = get_attendance_dir()
                    source_file = source_dir / f"{date_str}.json"
                    
                    checkin = 'Not recorded'
                    checkout = 'Not recorded'
                    delay = '00:00:00'
                    
                    if source_file.exists():
                        with source_file.open('r', encoding='utf-8') as f:
                            records = json.load(f)
                        
                        faculty_records = [r for r in records if r.get('student_id', '').strip().upper() == faculty_id.upper()]
                        
                        if faculty_records:
                            # Find earliest check-in and latest check-out
                            checkins = []
                            checkouts = []
                            
                            for record in faculty_records:
                                ci = parse_ts(record.get('checkin'))
                                co = parse_ts(record.get('checkout'))
                                if ci:
                                    checkins.append(ci)
                                if co:
                                    checkouts.append(co)
                            
                            if checkins:
                                earliest_checkin = min(checkins)
                                checkin = earliest_checkin.strftime('%H:%M:%S')
                            
                            if checkouts:
                                latest_checkout = max(checkouts)
                                checkout = latest_checkout.strftime('%H:%M:%S')
                            
                            # Calculate delay
                            staff_type = determine_staff_type(faculty_id)
                            delay_val = compute_daily_delay_for_records(faculty_records, date_obj, staff_type)
                            
                            # Check if both check-in and check-out are not recorded
                            if not checkins and not checkouts:
                                delay = 'Absent'
                            else:
                                delay = delay_val
                            
                            # Add to total delay
                            if delay_val and delay_val != 'N/A' and delay != 'Absent':
                                try:
                                    delay_parts = delay_val.split(':')
                                    if len(delay_parts) == 3:
                                        hours, minutes, seconds = map(int, delay_parts)
                                        total_delay_seconds += hours * 3600 + minutes * 60 + seconds
                                except:
                                    pass
                    
                    # Add row data
                    row_data = [
                        f"{day:02d}/{month.zfill(2)}/{year}",
                        checkin,
                        checkout,
                        delay
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = Border(
                            left=Side(style='thin', color='000000'),
                            right=Side(style='thin', color='000000'),
                            top=Side(style='thin', color='000000'),
                            bottom=Side(style='thin', color='000000')
                        )
                        
                        # Apply red text for "Absent" delay
                        if col == 4 and value == 'Absent':  # Delay column
                            cell.font = Font(color="FF0000", bold=True)  # Red text
                    
                    current_row += 1
                    
                except ValueError:
                    # Invalid date, skip
                    pass
            
            # Add total delay row
            total_delay_cell = ws.cell(row=current_row, column=1, value="Total Delay")
            total_delay_cell.font = Font(bold=True, size=12)
            total_delay_cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
            total_delay_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            total_delay_value = ws.cell(row=current_row, column=4, value=_seconds_to_hhmmss(total_delay_seconds))
            total_delay_value.font = Font(bold=True, size=12)
            total_delay_value.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
            total_delay_value.alignment = Alignment(horizontal="center", vertical="center")
            
            current_row += 2  # Add spacing between faculty members
        
        # Set column widths - increased for better visibility
        column_widths = [18, 20, 20, 18]  # Date, Check In, Check Out, Delay (increased from 15)
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # Save to BytesIO
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return app.response_class(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=faculty_detailed_report_{faculty}_{year}_{month}.xlsx'}
        )
        
    except Exception as e:
        return jsonify({"success": False, "error": f"Failed to export Excel: {str(e)}"})


@app.route("/api/faculty-detailed-report/pdf", methods=["GET"])
def api_faculty_detailed_report_pdf():
    """Export faculty detailed report as PDF file."""
    if not is_logged_in():
        return jsonify({"success": False, "error": "Not authenticated"}), 401
    
    try:
        month = request.args.get('month')
        year = request.args.get('year')
        faculty = request.args.get('faculty')
        
        if not month or not year or not faculty:
            return jsonify({"success": False, "error": "Month, year, and faculty are required"})
        
        # Get faculty details
        faculty_details = load_faculty_details()
        
        # Determine which faculty to process
        if faculty == 'all':
            target_faculty = list(faculty_details.keys())
        else:
            target_faculty = [faculty]
        
        # Get all days in the month
        from calendar import monthrange
        days_in_month = monthrange(int(year), int(month))[1]
        
        # Create PDF
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from io import BytesIO
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Get styles
        styles = getSampleStyleSheet()
        
        # Build content
        story = []
        
        # Use the header design
        header_table = create_pdf_header()
        story.append(header_table)
        story.append(Spacer(1, 20))
        
        # Convert month number to month name
        month_names = {
            '01': 'January', '02': 'February', '03': 'March', '04': 'April',
            '05': 'May', '06': 'June', '07': 'July', '08': 'August',
            '09': 'September', '10': 'October', '11': 'November', '12': 'December'
        }
        month_name = month_names.get(month.zfill(2), month)
        
        # Report title - Updated to match Excel format
        report_title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=12,
            alignment=1  # Center alignment
        )
        report_title = Paragraph(f"DETAILED FACULTY REPORT - {month_name}", report_title_style)
        story.append(report_title)
        story.append(Spacer(1, 20))
        
        # Process each faculty member
        for faculty_id in target_faculty:
            faculty_info = faculty_details.get(faculty_id, {})
            faculty_name = faculty_info.get('name', '')
            
            # Faculty header - Blue text format like second image
            faculty_header_style = ParagraphStyle(
                'FacultyHeader',
                parent=styles['Heading3'],
                fontSize=14,
                spaceAfter=10,
                textColor=colors.blue,  # Blue color like Excel
                fontName='Helvetica-Bold'
            )
            faculty_header = Paragraph(f"{faculty_id} - {faculty_name}", faculty_header_style)
            story.append(faculty_header)
            
            # Create table data for this faculty
            table_data = [['Date', 'Check In Time', 'Check Out Time', 'Delay']]
            total_delay_seconds = 0
            
            for day in range(1, days_in_month + 1):
                try:
                    date_str = f"{year}-{month.zfill(2)}-{day:02d}"
                    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                    
                    source_dir = get_attendance_dir()
                    source_file = source_dir / f"{date_str}.json"
                    
                    checkin = 'Not recorded'
                    checkout = 'Not recorded'
                    delay = '00:00:00'
                    
                    if source_file.exists():
                        with source_file.open('r', encoding='utf-8') as f:
                            records = json.load(f)
                        
                        faculty_records = [r for r in records if r.get('student_id', '').strip().upper() == faculty_id.upper()]
                        
                        if faculty_records:
                            # Find earliest check-in and latest check-out
                            checkins = []
                            checkouts = []
                            
                            for record in faculty_records:
                                ci = parse_ts(record.get('checkin'))
                                co = parse_ts(record.get('checkout'))
                                if ci:
                                    checkins.append(ci)
                                if co:
                                    checkouts.append(co)
                            
                            if checkins:
                                earliest_checkin = min(checkins)
                                checkin = earliest_checkin.strftime('%H:%M:%S')
                            
                            if checkouts:
                                latest_checkout = max(checkouts)
                                checkout = latest_checkout.strftime('%H:%M:%S')
                            
                            # Calculate delay
                            staff_type = determine_staff_type(faculty_id)
                            delay_val = compute_daily_delay_for_records(faculty_records, date_obj, staff_type)
                            
                            # Check if both check-in and check-out are not recorded
                            if not checkins and not checkouts:
                                delay = 'Absent'
                            else:
                                delay = delay_val
                            
                            # Add to total delay
                            if delay_val and delay_val != 'N/A' and delay != 'Absent':
                                try:
                                    delay_parts = delay_val.split(':')
                                    if len(delay_parts) == 3:
                                        hours, minutes, seconds = map(int, delay_parts)
                                        total_delay_seconds += hours * 3600 + minutes * 60 + seconds
                                except:
                                    pass
                    
                    # Add row data with special formatting for "Absent"
                    if delay == 'Absent':
                        # Use colored text for "Absent"
                        from reportlab.platypus import Paragraph
                        from reportlab.lib.styles import ParagraphStyle
                        absent_style = ParagraphStyle(
                            'AbsentStyle',
                            parent=styles['Normal'],
                            fontSize=10,
                            textColor=colors.red,
                            fontName='Helvetica-Bold',
                            alignment=1  # Center alignment
                        )
                        delay_para = Paragraph('<font color="red">Absent</font>', absent_style)
                        table_data.append([
                            f"{day:02d}/{month.zfill(2)}/{year}",
                            checkin,
                            checkout,
                            delay_para
                        ])
                    else:
                        table_data.append([
                            f"{day:02d}/{month.zfill(2)}/{year}",
                            checkin,
                            checkout,
                            delay
                        ])
                    
                except ValueError:
                    # Invalid date, skip
                    pass
            
            # Add total delay row
            table_data.append(['Total Delay', '', '', _seconds_to_hhmmss(total_delay_seconds)])
            
            # Create table with increased column widths
            table = Table(table_data, colWidths=[100, 100, 100, 100])
            table.setStyle(TableStyle([
                # Header styling with white borders
                ('BACKGROUND', (0, 0), (-1, 0), colors.black),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                # White borders for header columns
                ('LINEBEFORE', (0, 0), (0, 0), 2, colors.white),
                ('LINEAFTER', (0, 0), (0, 0), 2, colors.white),
                ('LINEBEFORE', (1, 0), (1, 0), 2, colors.white),
                ('LINEAFTER', (1, 0), (1, 0), 2, colors.white),
                ('LINEBEFORE', (2, 0), (2, 0), 2, colors.white),
                ('LINEAFTER', (2, 0), (2, 0), 2, colors.white),
                ('LINEBEFORE', (3, 0), (3, 0), 2, colors.white),
                ('LINEAFTER', (3, 0), (3, 0), 2, colors.white),
                # Data row styling
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 1), (-1, -1), 8),
                # Borders
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                # Total row styling - center aligned
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 11),
                ('ALIGN', (0, -1), (-1, -1), 'CENTER'),  # Center align total row
            ]))
            
            story.append(table)
            story.append(Spacer(1, 20))
            
            # Add page break if not the last faculty
            if faculty_id != target_faculty[-1]:
                story.append(PageBreak())
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        return app.response_class(
            buffer.getvalue(),
            mimetype='application/pdf',
            headers={'Content-Disposition': f'attachment; filename=faculty_detailed_report_{faculty}_{year}_{month}.pdf'}
        )
        
    except Exception as e:
        return jsonify({"success": False, "error": f"Failed to export PDF: {str(e)}"})


if __name__ == "__main__":
    print("Starting Faculty Attendance Admin System...")
    print(f"Reading attendance data directly from: {ATTENDANCE_DIR}")
    
    try:
        app.run(host='0.0.0.0', debug=True, port=7002)
    except Exception as e:
        print(f"Error starting server: {e}")


